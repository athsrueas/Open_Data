from __future__ import annotations

import csv
import json
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RAW_DATE = "2026-03-30"

WORLD_BANK_DIR = ROOT / "outputs" / "educational_inequality_map" / "phase1_world_bank"
RAW_DIR = ROOT / "outputs" / "educational_inequality_map" / "raw" / RAW_DATE
VIEWER_DIR = ROOT / "viewer" / "educational_inequality_map"
DATA_JS_PATH = VIEWER_DIR / "data.js"
ADULT_SKILLS_BACKFILL_PATH = ROOT / "outputs" / "educational_inequality_map" / "research" / "adult_skills_equivalent_backfill_seed.csv"
LITERACY_CAVEATS_PATH = ROOT / "outputs" / "educational_inequality_map" / "research" / "literacy_caveats.csv"
RESEARCH_DIR = ROOT / "outputs" / "educational_inequality_map" / "research"
MISSING_METRICS_CSV_PATH = RESEARCH_DIR / "viewer_missing_metrics_audit.csv"
MISSING_METRICS_MD_PATH = RESEARCH_DIR / "viewer_missing_metrics_audit.md"

WANTED_METRICS = {
    "adult_literacy_rate_15_plus": {
        "key": "literacyRate",
        "label": "Adult literacy",
        "category": "outcomes",
        "unit": "%",
        "higherIsBetter": True,
        "map": True,
    },
    "completion_rate_primary": {
        "key": "primaryCompletionRate",
        "label": "Primary completion",
        "category": "access",
        "unit": "%",
        "higherIsBetter": True,
        "map": True,
    },
    "government_expenditure_education_pct_gdp": {
        "key": "educationSpendPctGdp",
        "label": "Education spending",
        "category": "funding",
        "unit": "% of GDP",
        "higherIsBetter": True,
        "map": True,
    },
    "government_expenditure_education_pct_total_govt_expenditure": {
        "key": "educationSpendPctGovt",
        "label": "Education share of government spending",
        "category": "funding",
        "unit": "% of govt spend",
        "higherIsBetter": True,
        "map": False,
    },
    "gdp_per_capita_constant_usd": {
        "key": "gdpPerCapita",
        "label": "GDP per capita",
        "category": "cost",
        "unit": "USD",
        "higherIsBetter": True,
        "map": False,
    },
    "total_population": {
        "key": "population",
        "label": "Population",
        "category": "context",
        "unit": "people",
        "higherIsBetter": True,
        "map": False,
    },
}

MAP_METRICS = [
    "literacyRate",
    "primaryCompletionRate",
    "educationSpendPctGdp",
    "learningPovertyRate",
    "schoolAvailabilityPerMillion",
]


ALL_VIEWER_METRICS = [
    "literacyRate",
    "primaryCompletionRate",
    "educationSpendPctGdp",
    "educationSpendPctGovt",
    "gdpPerCapita",
    "population",
    "learningPovertyRate",
    "schoolAvailabilityTotal",
    "schoolAvailabilityPerMillion",
    "schoolsWithDataPct",
    "hloReadingScore",
]


def round_coordinates(value: Any, digits: int = 3) -> Any:
    if isinstance(value, list):
        return [round_coordinates(item, digits) for item in value]
    if isinstance(value, float):
        return round(value, digits)
    return value


def is_point(value: Any) -> bool:
    return (
        isinstance(value, list)
        and len(value) >= 2
        and all(isinstance(item, (int, float)) for item in value[:2])
    )


def thin_ring(ring: list[Any]) -> list[Any]:
    if len(ring) <= 80:
        return ring

    if len(ring) <= 180:
        step = 2
    elif len(ring) <= 500:
        step = 4
    else:
        step = 6

    thinned = [ring[0]] + ring[1:-1:step] + [ring[-1]]
    if len(thinned) < 4:
        return ring
    return thinned


def thin_coordinates(value: Any) -> Any:
    if not isinstance(value, list) or not value:
        return value
    if is_point(value[0]):
        return thin_ring(value)
    return [thin_coordinates(item) for item in value]


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def load_jurisdictions() -> dict[str, dict[str, Any]]:
    jurisdictions: dict[str, dict[str, Any]] = {}
    for row in read_csv_rows(WORLD_BANK_DIR / "jurisdictions.csv"):
        iso3 = row["iso3_code"]
        jurisdictions[iso3] = {
            "iso3": iso3,
            "name": row["name"],
            "region": row["world_bank_region"].strip(),
            "incomeGroup": row["world_bank_income_group"],
            "capital": row["capital_city"],
            "longitude": float(row["longitude"]) if row["longitude"] else None,
            "latitude": float(row["latitude"]) if row["latitude"] else None,
            "metrics": {},
            "sources": {},
        }
    return jurisdictions


def load_world_bank_metrics(countries: dict[str, dict[str, Any]]) -> None:
    latest_by_country_metric: dict[tuple[str, str], dict[str, Any]] = {}

    for row in read_csv_rows(WORLD_BANK_DIR / "observations.csv"):
        metric_id = row["metric_id"]
        if metric_id not in WANTED_METRICS:
            continue

        iso3 = row["source_country_code"]
        if iso3 not in countries:
            continue

        year = int(row["year"])
        key = (iso3, metric_id)
        existing = latest_by_country_metric.get(key)
        if existing is None or year > existing["year"]:
            latest_by_country_metric[key] = {
                "value": float(row["value_numeric"]) if row["value_numeric"] else None,
                "year": year,
                "sourceIndicatorCode": row["source_indicator_code"],
                "comparability": row["comparability_flag"],
                "quality": row["quality_flag"],
            }

    for (iso3, metric_id), payload in latest_by_country_metric.items():
        metric_meta = WANTED_METRICS[metric_id]
        countries[iso3]["metrics"][metric_meta["key"]] = {
            "value": payload["value"],
            "year": payload["year"],
            "label": metric_meta["label"],
            "unit": metric_meta["unit"],
        }
        countries[iso3]["sources"][metric_meta["key"]] = {
            "dataset": "World Bank EdStats",
            "indicator": payload["sourceIndicatorCode"],
            "comparability": payload["comparability"],
            "quality": payload["quality"],
        }


def load_us_piaac_literacy_equivalent(countries: dict[str, dict[str, Any]]) -> None:
    usa = countries.get("USA")
    if not usa:
        return

    if "literacyRate" in usa["metrics"]:
        return

    national_results_path = RAW_DIR / "nces_piaac_us" / "national_results_2023.html"
    if not national_results_path.exists():
        return

    html = national_results_path.read_text(encoding="utf-8", errors="ignore")
    normalized = re.sub(r"\s+", " ", html)
    match = re.search(r"(\d{1,3})\s*percent.*?Level 1 or below in literacy", normalized, re.IGNORECASE)
    if not match:
        return

    level_1_or_below = float(match.group(1))
    share_above_level_1 = 100.0 - level_1_or_below

    usa["metrics"]["literacyRate"] = {
        "value": share_above_level_1,
        "year": 2023,
        "label": "Adult literacy equivalent",
        "unit": "%",
    }
    usa["sources"]["literacyRate"] = {
        "dataset": "NCES PIAAC 2023",
        "indicator": "share_above_level_1_literacy",
        "comparability": "not_directly_comparable",
        "quality": "published_derived",
        "displayNote": f"U.S. fallback from NCES PIAAC 2023; derived as 100 minus {level_1_or_below:.0f} percent at Level 1 or below in literacy.",
    }


def load_adult_skills_equivalent_backfills(countries: dict[str, dict[str, Any]]) -> None:
    if not ADULT_SKILLS_BACKFILL_PATH.exists():
        return

    with ADULT_SKILLS_BACKFILL_PATH.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            iso3 = row["iso3"]
            country = countries.get(iso3)
            if not country or "literacyRate" in country["metrics"]:
                continue

            equivalent_pct = row.get("adult_literacy_equivalent_pct")
            if not equivalent_pct:
                continue

            country["metrics"]["literacyRate"] = {
                "value": float(equivalent_pct),
                "year": int(row["source_release_year"]) if row.get("source_release_year") else None,
                "label": "Adult literacy equivalent",
                "unit": "%",
            }
            coverage_note = row.get("coverage_note", "").strip()
            evidence_status = row.get("evidence_status", "").strip()
            note_parts = [
                f"Backfill from {row['source_system']}; derived as 100 minus {row['low_literacy_share_pct']} percent at Level 1 or below in literacy."
            ]
            if coverage_note:
                note_parts.append(coverage_note)
            if evidence_status:
                note_parts.append(evidence_status)

            country["sources"]["literacyRate"] = {
                "dataset": row["source_system"],
                "indicator": "share_above_level_1_literacy",
                "comparability": row.get("comparability_flag") or "not_directly_comparable",
                "quality": row.get("quality_flag") or "published_derived",
                "displayNote": " ".join(note_parts),
                "sourceUrl": row.get("source_url"),
            }


def apply_literacy_caveats(countries: dict[str, dict[str, Any]]) -> None:
    if not LITERACY_CAVEATS_PATH.exists():
        return

    with LITERACY_CAVEATS_PATH.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            iso3 = row.get("iso3", "").strip()
            metric_key = row.get("metric_key", "").strip()
            if not iso3 or not metric_key:
                continue

            country = countries.get(iso3)
            if not country:
                continue

            source = country["sources"].get(metric_key)
            if not source:
                continue

            display_note = row.get("display_note", "").strip()
            if display_note:
                existing = source.get("displayNote", "").strip()
                source["displayNote"] = f"{existing} {display_note}".strip() if existing else display_note

            caveat_source_url = row.get("caveat_source_url", "").strip()
            if caveat_source_url:
                source["caveatSourceUrl"] = caveat_source_url


def load_giga_metrics(countries: dict[str, dict[str, Any]]) -> None:
    rows = json.loads((RAW_DIR / "giga" / "api" / "locations_countries.json").read_text(encoding="utf-8"))
    for row in rows:
        iso3 = row.get("iso3_format")
        if iso3 not in countries:
            continue

        schools_total = row.get("schools_total")
        population_payload = countries[iso3]["metrics"].get("population")
        population = population_payload["value"] if population_payload else None
        schools_per_million = None
        if schools_total and population:
            schools_per_million = (float(schools_total) / float(population)) * 1_000_000

        countries[iso3]["metrics"]["schoolAvailabilityTotal"] = {
            "value": float(schools_total) if schools_total is not None else None,
            "year": None,
            "label": "Mapped schools",
            "unit": "schools",
        }
        countries[iso3]["metrics"]["schoolAvailabilityPerMillion"] = {
            "value": schools_per_million,
            "year": None,
            "label": "Schools per million residents",
            "unit": "schools / 1M people",
        }
        countries[iso3]["metrics"]["schoolsWithDataPct"] = {
            "value": float(row["schools_with_data_percentage"]) * 100 if row.get("schools_with_data_percentage") is not None else None,
            "year": None,
            "label": "Schools with data coverage",
            "unit": "%",
        }
        countries[iso3]["sources"]["schoolAvailabilityTotal"] = {
            "dataset": "Giga country metadata",
            "indicator": "schools_total",
            "comparability": "proxy",
            "quality": "experimental",
        }
        countries[iso3]["sources"]["schoolAvailabilityPerMillion"] = {
            "dataset": "Giga country metadata + World Bank population",
            "indicator": "schools_total / population",
            "comparability": "proxy",
            "quality": "experimental",
        }
        countries[iso3]["giga"] = {
            "connectivityAvailability": row.get("connectivity_availability"),
            "coverageAvailability": row.get("coverage_availability"),
            "schoolsWithDataPercentage": row.get("schools_with_data_percentage"),
        }


def load_learning_poverty(countries: dict[str, dict[str, Any]]) -> None:
    workbook = xlrd.open_workbook(
        str(RAW_DIR / "worldbank_research" / "downloads" / "lpv_edstats_1303_2024.xls"),
        on_demand=True,
    )
    sheet = workbook.sheet_by_name("WDI_indicators")
    latest: dict[str, dict[str, Any]] = {}

    for row_idx in range(1, sheet.nrows):
        row = sheet.row_values(row_idx)
        iso3, cty_or_agg, year, indicator, value, value_metadata, country_name, region_name = row[:8]
        if cty_or_agg != "cty" or indicator != "SE.LPV.PRIM":
            continue
        if iso3 not in countries or value in ("", None):
            continue

        record = latest.get(iso3)
        year_int = int(year)
        if record is None or year_int > record["year"]:
            latest[iso3] = {
                "value": float(value),
                "year": year_int,
                "metadata": value_metadata,
            }

    for iso3, payload in latest.items():
        countries[iso3]["metrics"]["learningPovertyRate"] = {
            "value": payload["value"],
            "year": payload["year"],
            "label": "Learning poverty",
            "unit": "%",
        }
        countries[iso3]["sources"]["learningPovertyRate"] = {
            "dataset": "World Bank Learning Poverty",
            "indicator": "SE.LPV.PRIM",
            "comparability": "partially_comparable",
            "quality": "published",
        }


def load_hlo(countries: dict[str, dict[str, Any]]) -> None:
    workbook = load_workbook(
        RAW_DIR / "worldbank_research" / "downloads" / "hlo_database.xlsx",
        read_only=True,
        data_only=True,
    )
    sheet = workbook["HLO Database"]
    latest: dict[str, dict[str, Any]] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        code, country, year, subject, level, sourcetest, n_res, hlo, hlo_se, hlo_m, hlo_m_se, hlo_f, hlo_f_se, region, incomegroup = row[:15]
        if code not in countries or subject != "reading" or hlo is None:
            continue

        score = 1 if level == "sec" else 0
        existing = latest.get(code)
        if existing is None or int(year) > existing["year"] or (int(year) == existing["year"] and score > existing["score"]):
            latest[code] = {
                "value": float(hlo),
                "year": int(year),
                "subject": subject,
                "level": level,
                "sourceTest": sourcetest,
                "score": score,
            }

    for iso3, payload in latest.items():
        countries[iso3]["metrics"]["hloReadingScore"] = {
            "value": payload["value"],
            "year": payload["year"],
            "label": "HLO reading",
            "unit": "score",
        }
        countries[iso3]["sources"]["hloReadingScore"] = {
            "dataset": "World Bank HLO",
            "indicator": f"reading_{payload['level']}",
            "comparability": "harmonized",
            "quality": "published",
            "sourceTest": payload["sourceTest"],
        }


def load_geojson(countries: dict[str, dict[str, Any]]) -> dict[str, Any]:
    metadata_rows = json.loads((RAW_DIR / "geoboundaries" / "api" / "gbOpen_ALL_ADM0.json").read_text(encoding="utf-8"))
    metadata_by_iso3 = {row["boundaryISO"]: row for row in metadata_rows}
    features: list[dict[str, Any]] = []

    for iso3, country in countries.items():
        geometry_path = RAW_DIR / "geoboundaries" / "downloads" / "ADM0_simplified" / f"{iso3}.geojson"
        if not geometry_path.exists():
            continue
        geojson = json.loads(geometry_path.read_text(encoding="utf-8"))
        feature = geojson["features"][0] if geojson.get("features") else geojson
        metadata = metadata_by_iso3.get(iso3, {})
        feature["geometry"]["coordinates"] = round_coordinates(
            thin_coordinates(feature["geometry"]["coordinates"])
        )
        feature["properties"] = {
            "iso3": iso3,
            "name": country["name"],
            "region": country["region"],
            "incomeGroup": country["incomeGroup"],
            "geometrySource": metadata.get("boundarySource"),
            "geometryYear": metadata.get("boundaryYearRepresented"),
        }
        features.append(feature)

    return {"type": "FeatureCollection", "features": features}


def build_country_records(countries: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    records = []
    for iso3, country in countries.items():
        records.append(
            {
                "iso3": iso3,
                "name": country["name"],
                "region": country["region"],
                "incomeGroup": country["incomeGroup"],
                "capital": country["capital"],
                "longitude": country["longitude"],
                "latitude": country["latitude"],
                "metrics": country["metrics"],
                "sources": country["sources"],
                "giga": country.get("giga", {}),
            }
        )
    return sorted(records, key=lambda row: row["name"])


def build_metric_catalog() -> list[dict[str, Any]]:
    catalog = []
    for meta in WANTED_METRICS.values():
        catalog.append(
            {
                "key": meta["key"],
                "label": meta["label"],
                "category": meta["category"],
                "unit": meta["unit"],
                "higherIsBetter": meta["higherIsBetter"],
                "map": meta["map"],
            }
        )

    catalog.extend(
        [
            {
                "key": "learningPovertyRate",
                "label": "Learning poverty",
                "category": "outcomes",
                "unit": "%",
                "higherIsBetter": False,
                "map": True,
            },
            {
                "key": "schoolAvailabilityTotal",
                "label": "Mapped schools",
                "category": "access",
                "unit": "schools",
                "higherIsBetter": True,
                "map": False,
            },
            {
                "key": "schoolAvailabilityPerMillion",
                "label": "Schools per million residents",
                "category": "access",
                "unit": "schools / 1M people",
                "higherIsBetter": True,
                "map": True,
            },
            {
                "key": "schoolsWithDataPct",
                "label": "Schools with data coverage",
                "category": "access",
                "unit": "%",
                "higherIsBetter": True,
                "map": False,
            },
            {
                "key": "hloReadingScore",
                "label": "HLO reading",
                "category": "outcomes",
                "unit": "score",
                "higherIsBetter": True,
                "map": False,
            },
        ]
    )
    return catalog


def build_data_model(metric_catalog: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "version": "2026-03-31",
        "primaryEntity": "country",
        "countryFields": [
            "iso3",
            "name",
            "region",
            "incomeGroup",
            "capital",
            "longitude",
            "latitude",
            "metrics",
            "sources",
            "giga",
        ],
        "metricPayloadFields": ["value", "year", "label", "unit"],
        "sourcePayloadFields": [
            "dataset",
            "indicator",
            "comparability",
            "quality",
            "displayNote",
            "sourceUrl",
            "caveatSourceUrl",
        ],
        "mapMetricKeys": MAP_METRICS,
        "metricKeys": [metric["key"] for metric in metric_catalog],
    }


def build_coverage_summary(countries: list[dict[str, Any]], metric_catalog: list[dict[str, Any]]) -> dict[str, Any]:
    metrics: dict[str, Any] = {}
    for metric in metric_catalog:
        key = metric["key"]
        with_data = 0
        with_notes = 0
        for country in countries:
            payload = country["metrics"].get(key)
            if payload and payload.get("value") is not None:
                with_data += 1
            source = country["sources"].get(key)
            if source and source.get("displayNote"):
                with_notes += 1

        metrics[key] = {
            "countriesWithData": with_data,
            "countriesMissingData": len(countries) - with_data,
            "countriesWithCaveatsOrEquivalents": with_notes,
        }

    return {"countryCount": len(countries), "metrics": metrics}


def build_missing_metrics_audit(countries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    audit_rows: list[dict[str, Any]] = []
    for country in countries:
        missing = [key for key in ALL_VIEWER_METRICS if country["metrics"].get(key, {}).get("value") is None]
        missing_map_metrics = [key for key in MAP_METRICS if country["metrics"].get(key, {}).get("value") is None]
        if not missing:
            continue

        population_payload = country["metrics"].get("population")
        population = population_payload.get("value") if population_payload else None
        audit_rows.append(
            {
                "iso3": country["iso3"],
                "name": country["name"],
                "region": country["region"],
                "income_group": country["incomeGroup"],
                "population": population,
                "missing_metric_count": len(missing),
                "missing_metrics": ";".join(missing),
                "missing_map_metric_count": len(missing_map_metrics),
                "missing_map_metrics": ";".join(missing_map_metrics),
            }
        )

    audit_rows.sort(
        key=lambda row: (
            -int(row["missing_map_metric_count"]),
            -(float(row["population"]) if row["population"] is not None else -1),
            row["name"],
        )
    )
    return audit_rows


def build_payload() -> dict[str, Any]:
    countries = load_jurisdictions()
    load_world_bank_metrics(countries)
    load_us_piaac_literacy_equivalent(countries)
    load_adult_skills_equivalent_backfills(countries)
    load_giga_metrics(countries)
    load_learning_poverty(countries)
    load_hlo(countries)
    apply_literacy_caveats(countries)
    country_records = build_country_records(countries)
    metric_catalog = build_metric_catalog()

    return {
        "generatedAt": datetime.now(UTC).isoformat(),
        "dataModel": build_data_model(metric_catalog),
        "coverageSummary": build_coverage_summary(country_records, metric_catalog),
        "title": "Educational Inequality Map",
        "visualizationGoals": [
            "Map education access disparities",
            "Compare countries on cost, access, funding, and outcomes",
            "Show how graph-style relationships can organize the story",
        ],
        "metricCatalog": metric_catalog,
        "mapMetricKeys": MAP_METRICS,
        "countries": country_records,
        "geojson": load_geojson(countries),
        "graphExamples": [
            {
                "from": "School availability proxy",
                "edge": "serves",
                "to": "Population",
                "note": "Giga school counts are currently used as a country-level access proxy.",
            },
            {
                "from": "Jurisdiction",
                "edge": "receives",
                "to": "Funding",
                "note": "World Bank education expenditure metrics show how much public resource reaches the system.",
            },
            {
                "from": "Jurisdiction",
                "edge": "produces",
                "to": "Outcomes",
                "note": "Literacy, completion, HLO, and learning poverty capture system performance from different angles.",
            },
        ],
        "sourceNotes": [
            "School availability is currently a proxy built from Giga mapped-school counts, not a universal official census.",
            "Funding and literacy are pulled from the latest available World Bank EdStats observations.",
            "The U.S. literacy fallback uses NCES PIAAC 2023 and is labeled as an equivalent, not a directly comparable adult literacy rate.",
            "Several additional high-income literacy gaps are backfilled from OECD adult-skills notes and remain labeled as not directly comparable equivalents.",
            "Learning poverty and HLO add an outcomes layer without forcing OECD PISA microdata into the first viewer.",
        ],
    }


def write_data_js(payload: dict[str, Any]) -> None:
    VIEWER_DIR.mkdir(parents=True, exist_ok=True)
    serialized = json.dumps(payload, separators=(",", ":"), ensure_ascii=True)
    DATA_JS_PATH.write_text(f"window.EDU_VIEWER_DATA = {serialized};\n", encoding="utf-8")


def write_missing_metrics_audit(payload: dict[str, Any]) -> None:
    RESEARCH_DIR.mkdir(parents=True, exist_ok=True)
    rows = build_missing_metrics_audit(payload["countries"])
    fieldnames = [
        "iso3",
        "name",
        "region",
        "income_group",
        "population",
        "missing_metric_count",
        "missing_metrics",
        "missing_map_metric_count",
        "missing_map_metrics",
    ]
    with MISSING_METRICS_CSV_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    top_rows = rows[:20]
    lines = [
        "# Viewer Missing Metrics Audit",
        "",
        f"Generated at: {payload['generatedAt']}",
        "",
        f"- Countries in payload: {len(payload['countries'])}",
        f"- Countries with at least one missing viewer metric: {len(rows)}",
        "",
        "## Highest-priority remaining gaps",
        "",
        "| Country | ISO3 | Missing map metrics | Missing all viewer metrics |",
        "|---|---|---:|---:|",
    ]
    for row in top_rows:
        lines.append(
            f"| {row['name']} | `{row['iso3']}` | {row['missing_map_metric_count']} | {row['missing_metric_count']} |"
        )
    lines.extend(
        [
            "",
            "See `viewer_missing_metrics_audit.csv` for the full country-level audit.",
            "",
        ]
    )
    MISSING_METRICS_MD_PATH.write_text("\n".join(lines), encoding="utf-8")


def run() -> None:
    payload = build_payload()
    write_data_js(payload)
    write_missing_metrics_audit(payload)
    print(f"Wrote viewer data to {DATA_JS_PATH.relative_to(ROOT)}")
    print(f"Wrote missing-metrics audit to {MISSING_METRICS_CSV_PATH.relative_to(ROOT)}")
    print(f"Countries: {len(payload['countries'])}")
    print(f"GeoJSON features: {len(payload['geojson']['features'])}")


if __name__ == "__main__":
    run()
